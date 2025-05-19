import PyPDF2  # Library to read PDF files
import re
import pandas as pd
import os
from openpyxl import load_workbook


def extract_text_from_pdf(pdf_file: str) -> [str]:
    """
    Extracts text from a PDF file and returns a list of strings,
    where each string represents the text on a page.
    """
    pdf_text = []
    try:
        with open(pdf_file, 'rb') as pdf:
            reader = PyPDF2.PdfReader(pdf, strict=False)
            for page in reader.pages:
                content = page.extract_text()
                if content:  # Ensure content is not None or empty
                    pdf_text.append(content)
    except Exception as e:
        print(f"Error reading PDF file: {e}")
    return pdf_text


def extract_shot_counts(text: str) -> tuple:
    """
    Extracts shot counts for home and away teams by finding the "Shots By Period" section
    and extracting the last two numbers from the next two lines.
    Returns a tuple (shots_home, shots_away).
    """
    # Find the "Shots By Period" section
    shots_section_start = text.find("Shots By Period")
    if shots_section_start == -1:
        print("Error: 'Shots By Period' section not found in the text.")
        return 0, 0

    # Extract the lines after "Shots By Period"
    shots_section = text[shots_section_start:]
    lines = shots_section.splitlines()

    # The next two lines after "Shots By Period" should contain the shot data
    if len(lines) < 3:
        print("Error: Not enough lines after 'Shots By Period'.")
        return 0, 0

    # Extract the last number from each line (total shots)
    try:
        American_shots = int(lines[1].split()[-1])  # Last number in the American line
        bucknell_shots = int(lines[2].split()[-1])  # Last number in the Bucknell line
    except (IndexError, ValueError) as e:
        print(f"Error parsing shot data: {e}")
        return 0, 0

    return American_shots, bucknell_shots

def extract_metadata(text: str) -> dict:
    """
    Extracts metadata (date, attendance, stadium, referee) from the text.
    Returns a dictionary with the extracted data.
    """
    metadata = {
        "date": None,
        "attendance": None,
        "stadium": None,
        "referee": None
    }

    # Extract date
    date_pattern = re.compile(r'Date:\s*(\d{1,2}/\d{1,2}/\d{4})')
    date_match = date_pattern.search(text)
    if date_match:
        metadata["date"] = date_match.group(1)

    # Extract attendance
    attendance_pattern = re.compile(r'Attendance:\s*(\d+)')
    attendance_match = attendance_pattern.search(text)
    if attendance_match:
        metadata["attendance"] = int(attendance_match.group(1))

    # Extract stadium
    stadium_pattern = re.compile(r'Stadium:\s*([\w\s]+)')
    stadium_match = stadium_pattern.search(text)
    if stadium_match:
        metadata["stadium"] = stadium_match.group(1).strip()

    # Extract referee (stops at "Asst" or "Asst.")
    referee_pattern = re.compile(r'Officials:\s*Referee,\s*([^,]+?)\s*(?:Asst\.?|$)')
    referee_match = referee_pattern.search(text)
    if referee_match:
        metadata["referee"] = referee_match.group(1).strip()

    return metadata

def extract_goals(text: str) -> tuple:
    """
    Extracts goal counts for home and away teams by finding the "Goals By Period" section
    and extracting the last two numbers from the next two lines.
    Returns a tuple (goals_home, goals_away).
    """
    # Find the "Goals By Period" section
    goals_section_start = text.find("Goals By Period")
    if goals_section_start == -1:
        print("Error: 'Goals By Period' section not found in the text.")
        return 0, 0

    # Extract the lines after "Goals By Period"
    goals_section = text[goals_section_start:]
    lines = goals_section.splitlines()

    # The next two lines after "Goals By Period" should contain the goal data
    if len(lines) < 3:
        print("Error: Not enough lines after 'Goals By Period'.")
        return 0, 0

    # Extract the last number from each line (total goals)
    try:
        American_goals = int(lines[1].split()[-1])  # Last number in the American line
        bucknell_goals = int(lines[2].split()[-1])  # Last number in the Bucknell line
    except (IndexError, ValueError) as e:
        print(f"Error parsing goal data: {e}")
        return 0, 0

    return American_goals, bucknell_goals

def extract_saves(text: str) -> tuple:
    """
    Extracts save counts for home and away teams by finding the "Saves By Period" section
    and extracting the third number from each of the next two lines.
    Returns a tuple (saves_home, saves_away).
    """
    # Find the "Saves By Period" section
    saves_section_start = text.find("Saves By Period")
    if saves_section_start == -1:
        print("Error: 'Saves By Period' section not found in the text.")
        return 0, 0

    # Extract the lines after "Saves By Period"
    saves_section = text[saves_section_start:]
    lines = saves_section.splitlines()

    # Check if there are at least three lines (for American and Bucknell)
    if len(lines) < 3:
        print("Error: Not enough lines after 'Saves By Period'.")
        return 0, 0

    # Check and extract the third number from each team's line
    American_line = lines[1].split()
    bucknell_line = lines[2].split()

    # Ensure each line has at least 3 numbers
    if len(American_line) < 3 or len(bucknell_line) < 3:
        print("Error: Unexpected format in 'Saves By Period'.")
        return 0, 0

    # Extract the third number (total saves)
    American_saves = (American_line[3])
    bucknell_saves = (bucknell_line[3][0])

    return American_saves, bucknell_saves


def extract_fouls(text: str) -> tuple:
    """
    Extracts foul counts for home and away teams by finding the "Fouls By Period" section
    and extracting the last two numbers from the next two lines.
    Returns a tuple (fouls_home, fouls_away).
    """
    # Find the "Fouls By Period" section
    fouls_section_start = text.find("Fouls By Period")
    if fouls_section_start == -1:
        print("Error: 'Fouls By Period' section not found in the text.")
        return 0, 0

    # Extract the lines after "Fouls By Period"
    fouls_section = text[fouls_section_start:]
    lines = fouls_section.splitlines()

    # The next two lines after "Fouls By Period" should contain the foul data
    if len(lines) < 3:
        print("Error: Not enough lines after 'Fouls By Period'.")
        return 0, 0

    # Extract the last number from each line (total fouls)
    try:
        American_fouls = int(lines[1].split()[-1])  # Last number in the American line
        bucknell_fouls = int(lines[2].split()[-1])  # Last number in the Bucknell line
    except (IndexError, ValueError) as e:
        print(f"Error parsing foul data: {e}")
        return 0, 0

    return American_fouls, bucknell_fouls

def extract_corner_kicks(text: str) -> tuple:
    """
    Extracts corner kick counts for home and away teams by finding the "Corner Kicks By Period" section
    and extracting the last two numbers from the next two lines.
    Returns a tuple (corner_kicks_home, corner_kicks_away).
    """
    # Find the "Corner Kicks By Period" section
    corner_kicks_section_start = text.find("Corner Kicks By Period")
    if corner_kicks_section_start == -1:
        print("Error: 'Corner Kicks By Period' section not found in the text.")
        return 0, 0

    # Extract the lines after "Corner Kicks By Period"
    corner_kicks_section = text[corner_kicks_section_start:]
    lines = corner_kicks_section.splitlines()

    # The next two lines after "Corner Kicks By Period" should contain the corner kick data
    if len(lines) < 3:
        print("Error: Not enough lines after 'Corner Kicks By Period'.")
        return 0, 0

    # Extract the last number from each line (total corner kicks)
    try:
        American_corner_kicks = int(lines[1].split()[-1])  # Last number in the American line
        bucknell_corner_kicks = int(lines[2].split()[-1])  # Last number in the Bucknell line
    except (IndexError, ValueError) as e:
        print(f"Error parsing corner kick data: {e}")
        return 0, 0

    return American_corner_kicks, bucknell_corner_kicks

def extract_home_and_away(text: str) -> tuple:
    """
    Extracts the home and away team names from the first line of the text.
    The away team is the first school, and the home team is the school after "-vs-".
    Returns a tuple (away_team, home_team).
    """
    # Match pattern "Team1 (...) -vs- Team2 (...)"
    match = re.match(r"(.+?)\s*\(.*?\)\s*-vs-\s*(.+?)\s*\(.*?\)", text.splitlines()[0])

    if match:
        away_team = match.group(1).strip()
        home_team = match.group(2).strip()
        return away_team, home_team
    else:
        print("Error: Unable to extract team names.")
        return None, None


def save_match_data_to_excel(metadata: dict, stats: dict, filename="match_data.xlsx"):
    """
    Saves the match metadata and statistics to an Excel file.
    Ensures that data is appended as a new row instead of replacing the existing data.
    """
    # Organize data to start with Home | Away | Date
    ordered_metadata = {
        "Home": metadata["home_team"],
        "Away": metadata["away_team"],
        "Date": metadata["date"],
        "Attendance": metadata["attendance"],
        "Stadium": metadata["stadium"],
        "Referee": metadata["referee"]
    }

    # Convert metadata and stats into a single DataFrame row
    match_data = {**ordered_metadata, **stats}
    df = pd.DataFrame([match_data])  # Convert dictionary to a DataFrame

    # Check if the file exists
    if os.path.exists(filename):
        with pd.ExcelWriter(filename, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            # Load existing workbook
            existing_book = load_workbook(filename)
            sheet = existing_book.active
            df.to_excel(writer, index=False, sheet_name="Match Stats", header=False, startrow=sheet.max_row)
    else:
        # If the file doesn't exist, create a new one with headers
        with pd.ExcelWriter(filename, mode='w', engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Match Stats")

    print(f"Match data successfully appended to {filename}!")




if __name__ == '__main__':
    # Extract text from the PDF
    extracted_text = extract_text_from_pdf("2024LehighvBostonUniversity.pdf")

    # Combine all pages into a single string for easier processing
    full_text = "\n".join(extracted_text)

    # Extract metadata and match statistics
    metadata = extract_metadata(full_text)
    shots_home, shots_away = extract_shot_counts(full_text)
    goals_home, goals_away = extract_goals(full_text)
    saves_home, saves_away = extract_saves(full_text)
    fouls_home, fouls_away = extract_fouls(full_text)
    corners_home, corners_away = extract_corner_kicks(full_text)
    away_team, home_team = extract_home_and_away(full_text)
    metadata["home_team"] = home_team
    metadata["away_team"] = away_team

    # Create a dictionary to store the match statistics
    stats_data = {
        "Statistic": ["Shots", "Goals", "Saves", "Fouls", "Corner Kicks"],
        "Colgate (Home)": [shots_home, goals_home, saves_home, fouls_home, corners_home],
        "Bucknell (Away)": [shots_away, goals_away, saves_away, fouls_away, corners_away],
    }

    # Convert the statistics dictionary into a pandas DataFrame
    stats_df = pd.DataFrame(stats_data)

    # Create a dictionary to store the metadata
    metadata_data = {
        "Metadata": ["Home Team", "Away Team", "Date", "Attendance", "Stadium", "Referee"],
        "Value": [home_team, away_team, metadata["date"], metadata["attendance"], metadata["stadium"], metadata["referee"]],
    }

    # Convert the metadata dictionary into a pandas DataFrame
    stats = {
        "home_team": home_team,
        "away_team": away_team,
        "shots_home": shots_home,
        "shots_away": shots_away,
        "goals_home": goals_home,
        "goals_away": goals_away,
        "saves_home": saves_home,
        "saves_away": saves_away,
        "fouls_home": fouls_home,
        "fouls_away": fouls_away,
        "corners_home": corners_home,
        "corners_away": corners_away
    }
    # Save the match data to an Excel file
    save_match_data_to_excel(metadata, stats)



